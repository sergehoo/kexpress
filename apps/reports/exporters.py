"""Génération des exports CSV / Excel / PDF à partir d'un dataset."""
from __future__ import annotations

import csv
import io


def to_csv(ds: dict) -> tuple[bytes, str]:
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")
    writer.writerow(ds["columns"])
    writer.writerows(ds["rows"])
    # BOM UTF-8 pour une ouverture correcte dans Excel.
    content = ("﻿" + buf.getvalue()).encode("utf-8")
    return content, "text/csv; charset=utf-8"


def to_xlsx(ds: dict) -> tuple[bytes, str]:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = ds["title"][:31]

    header_fill = PatternFill("solid", fgColor="1D3069")
    header_font = Font(bold=True, color="FFFFFF")
    for c, name in enumerate(ds["columns"], start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(ds["rows"], start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    # Largeur de colonnes auto (approchée)
    for c, name in enumerate(ds["columns"], start=1):
        width = max([len(str(name))] + [len(str(row[c - 1])) for row in ds["rows"]] + [8])
        ws.column_dimensions[get_column_letter(c)].width = min(width + 2, 40)
    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def to_pdf(ds: dict, subtitle: str = "") -> tuple[bytes, str]:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    out = io.BytesIO()
    doc = SimpleDocTemplate(out, pagesize=landscape(A4), title=ds["title"],
                            leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f"Kaydan Express — {ds['title']}", styles["Title"]),
    ]
    if subtitle:
        elements.append(Paragraph(subtitle, styles["Normal"]))
    elements.append(Spacer(1, 12))

    data = [ds["columns"]] + (ds["rows"] or [["Aucune donnée"] + [""] * (len(ds["columns"]) - 1)])
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D3069")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F6FB")]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E3E8F0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    doc.build(elements)
    return out.getvalue(), "application/pdf"
